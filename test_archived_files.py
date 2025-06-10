import os
import shutil
import zipfile
import pytest
import csv

from pypdf import PdfReader
from openpyxl import load_workbook

CURRENT_FILE = os.path.abspath(__file__)
CURRENT_DIR = os.path.dirname(CURRENT_FILE)
pdffile_withpath = os.path.abspath("tmp\\hw7.pdf")
xlsxfile_withpath = os.path.abspath("tmp\\hw7.xlsx")
csvfile_withpath = os.path.abspath("tmp\\hw7.csv")

shutil.copy2(pdffile_withpath, CURRENT_DIR)
shutil.copy2(xlsxfile_withpath, CURRENT_DIR)
shutil.copy2(csvfile_withpath, CURRENT_DIR)

@pytest.fixture
def archive_file():
    archname = "archive.zip"
    archfiles_list = [os.path.basename(pdffile_withpath),
                      os.path.basename(xlsxfile_withpath),
                      os.path.basename(csvfile_withpath)]

    if not os.path.exists(archname):
        with zipfile.ZipFile(archname, "w") as zf:
            for file in archfiles_list:
                zf.write(file)

def test_pdffile_archive(archive_file):
    reader = PdfReader(pdffile_withpath) # обращаемся к незапакованному pdf-файлу
    page = reader.pages[0]  # получаем первую страницу
    text = page.extract_text()  # извлекаем текст из первой страницы

    with zipfile.ZipFile('archive.zip') as zf:
        zipfile_list = (zf.namelist())
        with zf.open((zipfile_list[0])) as pdffile:
            zfreader = PdfReader(pdffile) # обращаемся к pdf-файлу в архиве
            zfpage = zfreader.pages[0]
            zftext = zfpage.extract_text()
    assert zftext == text

def test_xlsxfile_archive(archive_file):
    workbook = load_workbook(xlsxfile_withpath)  # обращаемся к незапакованному xlsx-файлу
    sheet = workbook.active  # получаем активный лист
    cell_value = sheet.cell(row=3, column=2).value # получаем значение в ячейке

    with zipfile.ZipFile('archive.zip') as zf:
        zipfile_list = (zf.namelist())
        with zf.open((zipfile_list[1])) as zfxlsxfile:
            zfworkbook = load_workbook(zfxlsxfile) # обращаемся к xlsx-файлу в архиве
            zfsheet = zfworkbook.active
            zfcell_value = zfsheet.cell(row=3, column=2).value
    assert zfcell_value == cell_value

def test_csvfile_archive(archive_file):
    with open(csvfile_withpath) as f:
        reader = csv.reader(f) # обращаемся к незапакованному csv-файлу
        readrow = [row for id, row in enumerate(reader) if id is 2] # читаем строку №2

    with zipfile.ZipFile('archive.zip') as zf:
        zipfile_list = (zf.namelist())
        with open(zipfile_list[2]) as csvfile:
            zfreader = csv.reader(csvfile)
            zfreadrow = [row for id, row in enumerate(zfreader) if id is 2]
    assert zfreadrow == readrow